import streamlit as st
import pandas as pd
import requests
import time
import os
import io
import sqlite3
import base64
import zipfile
from io import BytesIO
from datetime import datetime, timezone, timedelta
from requests.auth import HTTPBasicAuth
from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload, MediaIoBaseDownload
import plotly.express as px  # <-- LIBRERÍA PARA MEJORES GRÁFICOS
import openpyxl  # <-- PARA GENERAR EXCEL SOBRE LA PLANTILLA OFICIAL

# --- CONFIGURACIÓN DE PÁGINA ---
st.set_page_config(page_title="Extracción Energía", page_icon="🔌", layout="wide")

st.title("🔌 KERN IoP - Panel de Extracción Unificado (InfluxDB)")
st.write("Extrae consumos históricos desde Grafana/InfluxDB, analízalos y envíalos a la base de datos maestra.")

# --- CREDENCIALES Y CONSTANTES ---
USERNAME = "ahenao_estra"
PASSWORD = "X490fDvd"
AUTH = HTTPBasicAuth(USERNAME, PASSWORD)
HEADERS = {
    "Accept": "application/json",
    "Content-Type": "application/json"
}
URL_GRAFANA = "https://kern-iop.tech/api/ds/query"
TZ_COLOMBIA = timezone(timedelta(hours=-5))

# --- CONFIGURACIÓN DE ALMACENAMIENTO (Google Drive + SQLite) ---
FOLDER_ID_DRIVE = "131X02ZCk-UyABfxFMHZD0Loidb1jfIw3"
NOMBRE_DB = "energia.db"
RUTA_LOCAL_DB = "/tmp/energia.db"

# 🛠️ DICCIONARIO DE MAPEO INFLUXDB
MAPEO_INFLUX = {
    43: {"nombre": "S02", "tag_energia": "energy22", "tag_potencia": "energy22"},
    44: {"nombre": "H73", "tag_energia": "energy20", "tag_potencia": "energy20"},
    45: {"nombre": "H72", "tag_energia": "energy23", "tag_potencia": "energy23"},
    46: {"nombre": "H71", "tag_energia": "energy21", "tag_potencia": "energy21"},
    47: {"nombre": "H75", "tag_energia": "energy18", "tag_potencia": "energy18"},
    48: {"nombre": "H69", "tag_energia": "energy4", "tag_potencia": "energy4"},
    49: {"nombre": "H80", "tag_energia": "energy5", "tag_potencia": "energy5"},
    50: {"nombre": "H81", "tag_energia": "energy3", "tag_potencia": "energy3"},
    51: {"nombre": "H83", "tag_energia": "energy8", "tag_potencia": "energy8"},
    53: {"nombre": "H79", "tag_energia": "energy1", "tag_potencia": "energy1"},
    42: {"nombre": "H85", "tag_energia": "energy24", "tag_potencia": "energy24"},
    39: {"nombre": "H82", "tag_energia": "energy7", "tag_potencia": "energy7"},
    38: {"nombre": "H84", "tag_energia": "energy16", "tag_potencia": "energy16"},
    37: {"nombre": "H86", "tag_energia": "energy2", "tag_potencia": "energy2"},
    36: {"nombre": "H64", "tag_energia": "energy13", "tag_potencia": "energy13"},
    60: {"nombre": "H76", "tag_energia": "energy14", "tag_potencia": "energy14"},
    41: {"nombre": "H74", "tag_energia": "energy9", "tag_potencia": "energy9"}
}

# Generamos las opciones del select dinámicamente desde el diccionario
MAQUINAS_DISPONIBLES = {f"{v['nombre']} ({k})": k for k, v in MAPEO_INFLUX.items()}

# --- PLANTILLA OFICIAL "Plantilla_Energia-UTC.xlsx" EMBEBIDA (Base64) ---
# Se embebe para que la exportación por máquina no dependa de un archivo
# externo en el servidor de despliegue (Streamlit Cloud, Docker, etc.).
# Estructura de la plantilla (hoja "DATOS"):
#   Columna A -> "Demanda [kW]"
#   Columna B -> "Fecha [AAAA-MM-DDThh:mm:ss.dsZ] en formato UTC"
#   D1:E3     -> celdas de control (COUNTA/COUNT/resta), NO se tocan.
_PLANTILLA_XLSX_B64 = (
    "UEsDBBQABgAIAAAAIQCeLGxvawEAABAFAAATAAgCW0NvbnRlbnRfVHlwZXNdLnhtbCCiBAIooAACAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAACslMFOwzAMhu9IvEOVK2qzcUAIrdthwBEmMR4gJO4a"
    "LU2iOBvb2+NmY0KorELrpVEb+/+/uHYms11jsi0E1M6WbFyMWAZWOqXtqmTvy+f8nmUYhVXCOAsl2wOy2fT6arLce8CMsi2WrI7R"
    "P3COsoZGYOE8WNqpXGhEpNew4l7ItVgBvx2N7rh0NoKNeWw12HTyCJXYmJg97ejzgSSAQZbND4GtV8mE90ZLEYmUb6365ZIfHQrK"
    "TDFYa483hMF4p0O787fBMe+VShO0gmwhQnwRDWHwneGfLqw/nFsX50U6KF1VaQnKyU1DFSjQBxAKa4DYmCKtRSO0/eY+45+Ckadl"
    "PDBIe74k3MMR6X8DT8/LEZJMjyHGvQEcuuxJtM+5FgHUWww0GYMD/NTu4ZDCyHlNLTJwEU665/ypbxfBeaQJDvB/gO8RbbNzT0IQ"
    "oobTkHY1+8mRpv/iE0N7vyhQHd483WfTLwAAAP//AwBQSwMEFAAGAAgAAAAhALVVMCP0AAAATAIAAAsACAJfcmVscy8ucmVscyCi"
    "BAIooAACAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAACskk1PwzAM"
    "hu9I/IfI99XdkBBCS3dBSLshVH6ASdwPtY2jJBvdvyccEFQagwNHf71+/Mrb3TyN6sgh9uI0rIsSFDsjtnethpf6cXUHKiZylkZx"
    "rOHEEXbV9dX2mUdKeSh2vY8qq7iooUvJ3yNG0/FEsRDPLlcaCROlHIYWPZmBWsZNWd5i+K4B1UJT7a2GsLc3oOqTz5t/15am6Q0/"
    "iDlM7NKZFchzYmfZrnzIbCH1+RpVU2g5abBinnI6InlfZGzA80SbvxP9fC1OnMhSIjQS+DLPR8cloPV/WrQ08cudecQ3CcOryPDJ"
    "gosfqN4BAAD//wMAUEsDBBQABgAIAAAAIQCMOy4LSQMAAO8HAAAPAAAAeGwvd29ya2Jvb2sueG1spFXvb5s8EP4+af+DxXcKJuQH"
    "qOnUhLBVardozdovkSoXnGDV2Mw2Tapp//t7hpAuy6sp61BiY595eO7uueP8w7bk6JkqzaQYO/jMdxAVmcyZWI+db4vUHTlIGyJy"
    "wqWgY+eFaufDxft35xupnh6lfEIAIPTYKYypYs/TWUFLos9kRQVYVlKVxMBSrT1dKUpyXVBqSu4Fvj/wSsKE0yLE6hQMuVqxjCYy"
    "q0sqTAuiKCcG6OuCVbpDK7NT4EqinurKzWRZAcQj48y8NKAOKrP4ai2kIo8c3N7iPtoq+A3gj30Ygu5NYDp6VckyJbVcmTOA9lrS"
    "R/5j38P4IATb4xichhR6ij4zm8M9KzV4I6vBHmvwCob9f0bDIK1GKzEE741o/T23wLk4XzFO71rpIlJVn0lpM8UdxIk2s5wZmo+d"
    "ISzlhh5sqLqa1IyDNQhxDzvexV7Oc4VA/bTFWhRM3+90bg+BJi65oUoQQ6dSGJDgzqV/lVuDPS0kiBt9pd9rpijUFEgL3ISRZDF5"
    "1HNiClQrPnY+xstvguUkpxrogniVgZVeJlQTpSTnUqOcoltQ4IYoupwJo+jaSlmj+ezjl+WcE2EYh0ChjKg1QVCz7JksfxE1Oa6g"
    "v5A1yWzAPIhY61V7/3v0wDkVd9KdG4Xg/iq5hvTdkmdIJkgm39X6FWQL9x5EpmL88KM3iaJhECRuP/Qv3XDkD91RGoXuKAiimT/1"
    "J2Fv+BOcUYM4k6Q2xU4nFnrshCCKI9MN2XYW7Mc1y19p/PB3l2vn34bO9tM6bDviHaMb/aoou0TbeyZyuQHBjQYRtI6Xbu1iHzrs"
    "prHes9wUcCTyMZReu/eJsnUBlHF/ENl6VIGlNnYOKCUtpRQu1w4HlLxfODXNF7g1MxJNwSSXiy+30ORtX7ZBBnYqtq9QV3lTGl73"
    "VEZ4ZusDpiYbEfaDyHpNt+Zam2YGfTJgN+mPJn4vCtwwxakb4sh3J5NB6PaTtNcf4mQ666c2P/bbEW8t4uqNLWHkNU9TYmqoGVsu"
    "zTq2Y7rb3W+u2o2d5wdyjr8m1pXd0386eAvfRk5PPJzenXhw+vlmcXPi2evZ4uE+bQrsf731ICFQcF1avO5bffEfAAAA//8DAFBL"
    "AwQUAAYACAAAACEAkgeU7AQBAAA/AwAAGgAIAXhsL19yZWxzL3dvcmtib29rLnhtbC5yZWxzIKIEASigAAEAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAArJLLasQwDEX3hf6D0b5xMn1QhnFm0VKYbZt+gHCUOExiB1t95O9rUjrJwJBusjFIwvceibvbf3et"
    "+CQfGmcVZEkKgqx2ZWNrBe/Fy80jiMBoS2ydJQUDBdjn11e7V2qR46dgmj6IqGKDAsPcb6UM2lCHIXE92TipnO+QY+lr2aM+Yk1y"
    "k6YP0s81ID/TFIdSgT+UtyCKoY/O/2u7qmo0PTv90ZHlCxYy8NDGBUSBviZW8FsnkRHkZfvNmvYcz0KT+1jK8c2WGLI1Gb6cPwZD"
    "xBPHqRXkOFmEuV8TRmOrnww2doI5tZYucrdqKAx6Kt/Yx8zPszFv/8HIs9jnPwAAAP//AwBQSwMEFAAGAAgAAAAhAK2XVC7RAgAA"
    "xQYAABgAAAB4bC93b3Jrc2hlZXRzL3NoZWV0MS54bWycVV1vmzAUfZ+0/4D81D0EYyA0RSRV0qRaH9ZNW7s9O8YkXgEz2/mopv33"
    "XZuGpE1VRUWJv67vOef6XkN2ua1Kb82VFrIeIuIHyOM1k7moF0N0f3fdGyBPG1rntJQ1H6JHrtHl6OOHbCPVg15ybjxAqPUQLY1p"
    "Uow1W/KKal82vAZLIVVFDUzVAutGcZo7p6rEYRAkuKKiRi1Cqk7BkEUhGJ9Ktqp4bVoQxUtqQL9eikbv0Cp2ClxF1cOq6TFZNQAx"
    "F6Uwjw4UeRVLbxa1VHReQtxbElPmbRX8QvhHOxq3fsRUCaakloXxARm3mo/Dv8AXmLIO6Tj+k2BIjBVfC5vAPVT4Pkmk32GFe7Do"
    "nWBJB2aPS6UrkQ/R3+Dp6UFPbBP0AmKbg+cfGmWuTr4pD4qR39IKcvBZ/qYhwqMsF5B8G7CneDFEY5LOIrvuXH4KvtEHY8/Q+Q9e"
    "cmY40BPk2cqdS/lgN97AUmDJ3AaLSJkRa37Fy3KIJnAI+o/jgCEQ4I7hcLxju3a1DpLnVPMrWf4SuVkCJdypnBd0VZr94sAfDJJz"
    "Mjjvd8bvcvOZi8XSgEvsx3BktubS/HHKNYNiB6W+C5PJEiih9SphLy3UKt22obWMIehmK21ktZNgxXcOYHUO0G9ah7jvR1EYRCQE"
    "OXOuzbWwKt5EgZw6FOifUMi53+/HiYvpDXqIzDlCv3OM/SSJg8SyHzviNl539FNq6ChTcuPBbQF9uqH23UNS8Hz1uCBsu3UMeyEi"
    "Dalej4IMryF/7Mk2ObTFz23TQxt5bpsRQCtGV1/vb+/GZ+MwHfcTW9OfMlw8p8EguFNta+pE1VPY26kOX7CHe/azSZhOTiKHZJ1M"
    "Dns78ugFeeTIZ6Q3A1mvRttelTZfDV3wL1QtRK29kheukm3hq7baA99OjGxcSufSQOG64RI+FxyyF/hgL6Q0u4m9it0HaPQfAAD/"
    "/wMAUEsDBBQABgAIAAAAIQDppiW4ZgYAAFMbAAATAAAAeGwvdGhlbWUvdGhlbWUxLnhtbOxZzW4bNxC+F+g7EHtPLNmSYhmRA0uW"
    "4jZxYthKihypXWqXEXe5ICk7uhXJsUCBomnRS4HeeijaBkiAXtKncZuiTYG8QofkSlpaVGwnBvoXHWwt9+P8z3CGunrtQcrQIRGS"
    "8qwVVC9XAkSykEc0i1vBnX7v0nqApMJZhBnPSCuYEBlc23z/vat4QyUkJQj2Z3IDt4JEqXxjZUWGsIzlZZ6TDN4NuUixgkcRr0QC"
    "HwHdlK2sViqNlRTTLEAZToHs7eGQhgT1Nclgc0q8y+AxU1IvhEwcaNLE2WGw0aiqEXIiO0ygQ8xaAfCJ+FGfPFABYlgqeNEKKuYT"
    "rGxeXcEbxSamluwt7euZT7Gv2BCNVg1PEQ9mTKu9WvPK9oy+ATC1iOt2u51udUbPAHAYgqZWljLNWm+92p7SLIHs10XanUq9UnPx"
    "JfprCzI32+12vVnIYokakP1aW8CvVxq1rVUHb0AWX1/A19pbnU7DwRuQxTcW8L0rzUbNxRtQwmg2WkBrh/Z6BfUZZMjZjhe+DvD1"
    "SgGfoyAaZtGlWQx5ppbFWorvc9EDgAYyrGiG1CQnQxxCFHdwOhAUawZ4g+DSG7sUyoUlzQvJUNBctYIPcwwZMaf36vn3r54/Ra+e"
    "Pzl++Oz44U/Hjx4dP/zR0nI27uAsLm98+e1nf379Mfrj6TcvH3/hx8sy/tcfPvnl58/9QMiguUQvvnzy27MnL7769PfvHnvgWwIP"
    "yvA+TYlEt8gR2ucp6GYM40pOBuJ8O/oJps4OnABtD+muShzgrQlmPlybuMa7K6B4+IDXx/cdWQ8SMVbUw/lGkjrAXc5ZmwuvAW5o"
    "XiUL98dZ7GcuxmXcPsaHPt4dnDmu7Y5zqJrToHRs30mII+Yew5nCMcmIQvodHxHi0e4epY5dd2kouORDhe5R1MbUa5I+HTiBNN+0"
    "Q1Pwy8SnM7jasc3uXdTmzKf1Njl0kZAQmHmE7xPmmPE6Hiuc+kj2ccrKBr+JVeIT8mAiwjKuKxV4OiaMo25EpPTtuS1A35LTb2Co"
    "V16377JJ6iKFoiMfzZuY8zJym486CU5zr8w0S8rYD+QIQhSjPa588F3uZoh+Bj/gbKm771LiuPv0QnCHxo5I8wDRb8aiqNpO/U1p"
    "9rpizChU43fFeHo6bcHR5EuJnRMleBnuX1h4t/E42yMQ64sHz7u6+67uBv/5urssl89abecFFprkeV9suuR0aZM8pIwdqAkjN6Xp"
    "kyUcFlEPFk0Db6a42dCUJ/C1KO4OLhbY7EGCq4+oSg4SnEOPXTUjXywL0rFEOZcw25llM3ySE7TNOEmhzTaTYV3PDLYeSKx2eWSX"
    "18qz4YyMmRRjM39OGa1pAmdltnbl7ZhVrVRLzeaqVjWimVLnqDZTGXy4qBoszqwJXQiC3gWs3IARXcsOswlmJNJ2t3Pz1C2a9YW6"
    "SCY4IoWPtN6LPqoaJ01jZRpGHh/pOe8UH5W4NTXZt+B2FieV2dWWsJt67228NB1u517SeXsiHVlWTk6WoaNW0Kyv1gMU4rwVDGGs"
    "ha9pDl6XuvHDLIa7oVAJG/anJrMJ17k3m/6wrMJNhbX7gsJOHciFVNtYJjY0zKsiBFhmhnAj/2odzHpRCthIfwMp1tYhGP42KcCO"
    "rmvJcEhCVXZ2acXcURhAUUr5WBFxkERHaMDGYh+D+3Wogj4RlXA7YSqCfoCrNG1t88otzkXSlS+wDM6uY5YnuCi3OkWnmWzhJo9n"
    "MpgnK60RD3Tzym6UO78qJuUvSJVyGP/PVNHnCVwXrEXaAyHc5AqMdL62Ai5UwqEK5QkNewIuuUztgGiB61h4DUEF98nmvyCH+r/N"
    "OUvDpDVMfWqfxkhQOI9UIgjZg7Jkou8UYtXi7LIkWUHIRFRJXJlbsQfkkLC+roENfbYHKIFQN9WkKAMGdzL+3OcigwaxbnL+qZ2P"
    "Tebztge6O7Atlt1/xl6kVir6paOg6T37TE81KwevOdjPedTairWg8Wr9zEdtDpc+SP+B84+KkNkfJ/SB2uf7UFsR/NZg2ysEUX3J"
    "Nh5IF0hbHgfQONlFG0yalG1Yiu72wtsouJEuOt0ZX8jSN+l0z2nsWXPmsnNy8fXd5/mMXVjYsXW50/WYGpL2ZIrq9mg6yBjHmF+1"
    "yj888cF9cPQ2XPGPmZL2av8BXPHBlGF/JIDkt841Wzf/AgAA//8DAFBLAwQUAAYACAAAACEA3Lr+p8QCAAAwBwAADQAAAHhsL3N0"
    "eWxlcy54bWy0Vd9vmzAQfp+0/8HyOzWQwJIIqJamSJW2aVI7aa8GTGLVPyLjZGTT/vedgSRErbap3XgI5+P83Xf32ZfkupUC7Zlp"
    "uFYpDq58jJgqdcXVOsVfHnJvhlFjqaqo0Iql+MAafJ29fZM09iDY/YYxiwBCNSneWLtdENKUGyZpc6W3TMGXWhtJLSzNmjRbw2jV"
    "uE1SkND3YyIpV7hHWMjyb0AkNY+7rVdquaWWF1xwe+iwMJLl4m6ttKGFAKptMKUlaoPYhKg1xySd90keyUujG13bK8Aluq55yZ7S"
    "nZM5oeUZCZBfhhRExA8vam/NC5GmxLA9d/LhLKm1sg0q9U7ZFE+AqGvB4lHpbyp3n0DhISpLmu9oTwV4AkyypNRCG2RBOuhc51FU"
    "sj7ihgpeGO7Caiq5OPTu0Dk6tYc4yaH3zkkcj57N83nMukhxDo8Pj9vxr5IVjtIzhf2HhF2RDVTJhTj1PHTtBUeWwOG0zKgcFmiw"
    "Hw5baK6Ce9Q3qYv7Q/Ta0EMQRqMNpEuYJYU2Fdzbo9pO2N6VJYLVFhph+Hrj3lZv4bfQ1sLZzpKK07VWVDihjjsGA8opmRD37m5/"
    "rS+w2xqpncylvatSDFPCSXw0oZDB7PH6hcMfo/XYr4ZFbX2JD4gj2hekT+mRO2Ap/uSGkYB7MUCgYseF5eoZwoBZtecWdKfUusHS"
    "NeeUBTpRsZruhH04fUzx2f7IKr6T81PUZ77XtoNI8dn+4JQKYqcya+2HBi4PvNHO8BT/uF2+m69u89Cb+cuZN52wyJtHy5UXTW+W"
    "q1U+90P/5udovL1iuHXTOEtgbCwaASPQDMUOJd6ffSkeLXr63RkF2mPu8zD230eB7+UTP/CmMZ15s3gSeXkUhKt4uryN8mjEPXrh"
    "EPRJEPTj1JGPFpZLJrg6anVUaOwFkWD5myLIUQly/qvLfgEAAP//AwBQSwMEFAAGAAgAAAAhAJVd0h9mAQAAzQMAABQAAAB4bC9z"
    "aGFyZWRTdHJpbmdzLnhtbMRTXUvDMBR9F/wPIe9buoEipc2Qjb4NfNgQHHvI2qwN5qPmZsX5b/wt/jFvWhVkPvgg7lIKOSc35+bk"
    "3mz2bDTppAflbE4n44QSaUtXKVvndL0qRjeUQBC2EtpZmdOjBDrjlxcZQCCYayGnTQhtyhiUjTQCxq6VFpm980YEXPqaQeulqKCR"
    "MhjNpklyzYxQlpLSHWzI6RUlB6ueDnL+ueYZKJ4FvsAjUZxsHu+3GQs8YxEfuEJpAaSVtagEnJCybJDt3l61+onucwXxslMg/Pds"
    "H5WLeADZ3GKMlsvRYjHsQS5+d/iDF9IJja5NKONZ6bTzxNe7nBYYCUaEfeFsGPbNhVY7ryK6F0bp4wBPI9CbJwfAKOt8BFkvE/jq"
    "V9IB7ccX6ov5C9WmSY1JAcbVh7nnuPrDOa4eGzuFVpRoJ3YuSN9Jyrc4GeSknF18vP9qhGGmHFmv5l+F9APBcBz5OwAAAP//AwBQ"
    "SwMEFAAGAAgAAAAhAOgpiuWiAAAA0QAAABAAAAB4bC9jYWxjQ2hhaW4ueG1sPI7BCsIwEETvgv8Q9m7TVhCRpj0U/QL9gJCuTSDZ"
    "lGwQ/XuD2F4G9u0wM93wDl68MLGLpKCpahBIJk6OZgWP++1wBsFZ06R9JFTwQYah3+86o70ZrXYkSgKxApvzcpGSjcWguYoLUvk8"
    "Ywo6lzPNkpeEemKLmIOXbV2fZCgB0HdGJAXXFoRTUNSXJSBX3PzxBo6rr7T+fHIb038BAAD//wMAUEsDBBQABgAIAAAAIQDO4+d6"
    "SgEAAF0CAAARAAgBZG9jUHJvcHMvY29yZS54bWwgogQBKKAAAQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAACMks1qwzAQ"
    "hO+FvoPR3ZblpCEI26E/5NRAIQ4tvQlpk5haspDUOn77ynbiOrSHHrUz++3sonR1klXwBcaWtcoQiWIUgOK1KNUhQ7tiHS5RYB1T"
    "glW1ggy1YNEqv71Juaa8NvBiag3GlWADT1KWcp2ho3OaYmz5ESSzkXcoL+5rI5nzT3PAmvEPdgCcxPECS3BMMMdwBwz1SERnpOAj"
    "Un+aqgcIjqECCcpZTCKCf7wOjLR/NvTKxClL12q/0znulC34II7uky1HY9M0UTPrY/j8BL9tnrf9qmGpultxQHkqOOUGmKtN3u2v"
    "21OV4kmxO2DFrNv4W+9LEA9tfr/dbVP8u+5ZffQBCCLwYegQ/aK8zh6fijXKkzhZhPEyjOcFITRZ0nj+3o296u/CDQV5Hv4f4qIg"
    "M5oQepdMiBdA3ue+/hD5NwAAAP//AwBQSwMEFAAGAAgAAAAhAAvZpn2LAQAABAMAABAACAFkb2NQcm9wcy9hcHAueG1sIKIEASig"
    "AAEAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAnJJBbtswEEX3BXoHgvuYcloEhUExCOIWXrSIASvZs9TIZkuTAmcs2L1N"
    "z9KLdSQhjpx01d3M/NHX4yf17XEfRAcZfYqlnM8KKSC6VPu4LeVj9eXqkxRINtY2pAilPAHKW/P+nV7n1EImDyjYImIpd0TtQil0"
    "O9hbnLEcWWlS3lviNm9VahrvYJncYQ+R1HVR3Cg4EsQa6qv2bChHx0VH/2taJ9fz4VN1ahnY6Lu2Dd5Z4lOab97lhKkh8fnoIGg1"
    "FTXTbcAdsqeTKbSatnrjbIB7NjaNDQhavQz0Cmwf2tr6jEZ3tOjAUcoC/S+O7VqK7xahxyllZ7O3kRirXxuboQ4tUjar9MOiqEG4"
    "P7+DO4SkFe+N2lBOP5nW/qOZDwtcXC72BiMPC5eklacA+NCsbaZ/gM+n4APDiD3iLO+qh80bvOHg/KNX1l99/ImPbZWWluA5wcuh"
    "3uxshppDPyd8HugVh5dDb3K/s3EL9fPOW6G/76fxUZv5zaz4UPBVTmZavTxf8xcAAP//AwBQSwECLQAUAAYACAAAACEAnixsb2sB"
    "AAAQBQAAEwAAAAAAAAAAAAAAAAAAAAAAW0NvbnRlbnRfVHlwZXNdLnhtbFBLAQItABQABgAIAAAAIQC1VTAj9AAAAEwCAAALAAAA"
    "AAAAAAAAAAAAAKQDAABfcmVscy8ucmVsc1BLAQItABQABgAIAAAAIQCMOy4LSQMAAO8HAAAPAAAAAAAAAAAAAAAAAMkGAAB4bC93"
    "b3JrYm9vay54bWxQSwECLQAUAAYACAAAACEAkgeU7AQBAAA/AwAAGgAAAAAAAAAAAAAAAAA/CgAAeGwvX3JlbHMvd29ya2Jvb2su"
    "eG1sLnJlbHNQSwECLQAUAAYACAAAACEArZdULtECAADFBgAAGAAAAAAAAAAAAAAAAACDDAAAeGwvd29ya3NoZWV0cy9zaGVldDEu"
    "eG1sUEsBAi0AFAAGAAgAAAAhAOmmJbhmBgAAUxsAABMAAAAAAAAAAAAAAAAAig8AAHhsL3RoZW1lL3RoZW1lMS54bWxQSwECLQAU"
    "AAYACAAAACEA3Lr+p8QCAAAwBwAADQAAAAAAAAAAAAAAAAAhFgAAeGwvc3R5bGVzLnhtbFBLAQItABQABgAIAAAAIQCVXdIfZgEA"
    "AM0DAAAUAAAAAAAAAAAAAAAAABAZAAB4bC9zaGFyZWRTdHJpbmdzLnhtbFBLAQItABQABgAIAAAAIQDoKYrlogAAANEAAAAQAAAA"
    "AAAAAAAAAAAAAKgaAAB4bC9jYWxjQ2hhaW4ueG1sUEsBAi0AFAAGAAgAAAAhAM7j53pKAQAAXQIAABEAAAAAAAAAAAAAAAAAeBsA"
    "AGRvY1Byb3BzL2NvcmUueG1sUEsBAi0AFAAGAAgAAAAhAAvZpn2LAQAABAMAABAAAAAAAAAAAAAAAAAA+R0AAGRvY1Byb3BzL2Fw"
    "cC54bWxQSwUGAAAAAAsACwC+AgAAuiAAAAAA"
)


def _cargar_plantilla_bytes() -> BytesIO:
    """Decodifica la plantilla oficial embebida y la devuelve lista para abrir con openpyxl."""
    return BytesIO(base64.b64decode(_PLANTILLA_XLSX_B64))


def _timestamp_bogota_a_utc_iso(timestamp_bogota_str: str) -> str:
    """Convierte un Timestamp guardado en hora Bogotá (naive, 'YYYY-MM-DD HH:MM:SS')
    al formato UTC ISO exigido por la plantilla: 'AAAA-MM-DDThh:mm:ss.dsZ'.
    Como la fuente no trae sub-segundos, el decisegundo se fija en 0.
    Se usa para el CSV, donde la fecha viaja como texto plano."""
    dt_bogota = pd.to_datetime(timestamp_bogota_str).tz_localize(TZ_COLOMBIA)
    dt_utc = dt_bogota.astimezone(timezone.utc)
    return dt_utc.strftime("%Y-%m-%dT%H:%M:%S.0Z")


def _timestamp_bogota_a_utc_dt(timestamp_bogota_str: str) -> datetime:
    """Igual que _timestamp_bogota_a_utc_iso pero devuelve un datetime real (naive,
    sin tzinfo) en vez de texto. Se usa para el Excel: la columna B de la plantilla
    valida las fechas con =COUNT(B2:B560001), y COUNT solo cuenta valores numéricos/
    de fecha reales, NO texto. Si se escribe un string (aunque tenga forma de fecha),
    Excel lo trata como texto y 'Fechas válidas' siempre da 0."""
    dt_bogota = pd.to_datetime(timestamp_bogota_str).tz_localize(TZ_COLOMBIA)
    dt_utc = dt_bogota.astimezone(timezone.utc)
    return dt_utc.replace(tzinfo=None)


# Formato de celda personalizado: la celda guarda una fecha REAL (para que
# COUNT() la cuente como válida), pero se ve idéntica al formato pedido
# "AAAA-MM-DDThh:mm:ss.dsZ" (el ".0Z" final se escribe como texto literal,
# ya que la fuente no trae sub-segundos reales).
_FORMATO_FECHA_UTC = 'yyyy-mm-dd"T"hh:mm:ss".0Z"'


def generar_csv_maquina(df_maquina: pd.DataFrame) -> bytes:
    """Genera el contenido CSV de UNA máquina con las columnas exactas de la plantilla."""
    df_out = pd.DataFrame({
        "Demanda [kW]": df_maquina["Potencia_kW"].values,
        "Fecha [AAAA-MM-DDThh:mm:ss.dsZ] en formato UTC": df_maquina["Fecha_UTC"].values,
    })
    return df_out.to_csv(index=False, encoding="utf-8-sig").encode("utf-8-sig")


def generar_xlsx_maquina(df_maquina: pd.DataFrame) -> bytes:
    """Genera el .xlsx de UNA máquina reutilizando la plantilla oficial tal cual
    (mismos encabezados, mismas fórmulas de control en D1:E3). Solo se escriben
    los datos a partir de la fila 2 en las columnas A y B. La fecha se escribe
    como valor de fecha real (no texto) para que =COUNT(...) la reconozca."""
    wb = openpyxl.load_workbook(_cargar_plantilla_bytes())
    ws = wb["DATOS"]

    fila = 2
    for demanda, fecha_utc_dt in zip(df_maquina["Potencia_kW"].values, df_maquina["Fecha_UTC_dt"].values):
        ws.cell(row=fila, column=1, value=float(demanda))
        celda_fecha = ws.cell(row=fila, column=2, value=pd.Timestamp(fecha_utc_dt).to_pydatetime())
        celda_fecha.number_format = _FORMATO_FECHA_UTC
        fila += 1

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# --- FUNCIONES DE CONEXIÓN A GOOGLE DRIVE ---
@st.cache_resource
def conectar_drive():
    """Autoriza el service account con scope de Drive y devuelve el cliente de la API."""
    scope = ["https://www.googleapis.com/auth/drive"]
    creds_dict = st.secrets["gcp_service_account"]
    creds = Credentials.from_service_account_info(creds_dict, scopes=scope)
    return build("drive", "v3", credentials=creds)


def buscar_db_en_drive(drive_service):
    """Busca energia.db dentro de la carpeta de Drive. Devuelve el file_id o None si no existe."""
    query = f"name='{NOMBRE_DB}' and '{FOLDER_ID_DRIVE}' in parents and trashed=false"
    resultados = drive_service.files().list(q=query, fields="files(id, name)").execute()
    archivos = resultados.get("files", [])
    return archivos[0]["id"] if archivos else None


def descargar_db(drive_service, file_id):
    """Descarga el archivo energia.db de Drive a /tmp."""
    request = drive_service.files().get_media(fileId=file_id)
    with io.FileIO(RUTA_LOCAL_DB, "wb") as fh:
        downloader = MediaIoBaseDownload(fh, request)
        done = False
        while not done:
            _, done = downloader.next_chunk()


def subir_db(drive_service, file_id=None):
    """Sube (o actualiza) el archivo energia.db en Drive. Devuelve el file_id."""
    media = MediaFileUpload(RUTA_LOCAL_DB, mimetype="application/x-sqlite3")
    if file_id:
        drive_service.files().update(fileId=file_id, media_body=media).execute()
        return file_id
    else:
        metadata = {"name": NOMBRE_DB, "parents": [FOLDER_ID_DRIVE]}
        nuevo = drive_service.files().create(body=metadata, media_body=media, fields="id").execute()
        return nuevo["id"]


def preparar_tabla(conn):
    """Crea la tabla de registros si aún no existe, con clave única para evitar duplicados."""
    conn.execute("""
        CREATE TABLE IF NOT EXISTS registros_energia (
            Timestamp TEXT NOT NULL,
            ID_Maquina_Texto TEXT,
            id_maquina_api INTEGER,
            Energia_kWh REAL,
            Potencia_kW REAL,
            UNIQUE(Timestamp, id_maquina_api)
        )
    """)
    conn.commit()


def upsertar_dataframe(conn, df):
    """Inserta los registros nuevos e ignora silenciosamente los que ya existan (misma clave)."""
    registros = df[["Timestamp", "ID_Maquina_Texto", "id_maquina_api", "Energia_kWh", "Potencia_kW"]].values.tolist()
    conn.executemany(
        """INSERT OR IGNORE INTO registros_energia
           (Timestamp, ID_Maquina_Texto, id_maquina_api, Energia_kWh, Potencia_kW)
           VALUES (?, ?, ?, ?, ?)""",
        registros
    )
    conn.commit()


# --- FUNCIONES DE EXTRACCIÓN GRAFANA / INFLUXDB ---
def extraer_dataframe_json(json_resp, ref_id, nombre_columna):
    try:
        valores = json_resp['results'][ref_id]['frames'][0]['data']['values']
        return pd.DataFrame({'ts': valores[0], nombre_columna: valores[1]})
    except (KeyError, IndexError, TypeError):
        return pd.DataFrame(columns=['ts', nombre_columna])

def api_post_grafana_energia(dt_inicio_aware, dt_fin_aware, tag_energia):
    dt_inicio_utc = dt_inicio_aware.astimezone(timezone.utc)
    dt_fin_utc = dt_fin_aware.astimezone(timezone.utc)

    iso_inicio = dt_inicio_utc.strftime('%Y-%m-%dT%H:%M:%S.000Z')
    iso_fin = dt_fin_utc.strftime('%Y-%m-%dT%H:%M:%S.000Z')
    ts_inicio_ms = str(int(dt_inicio_aware.timestamp() * 1000))
    ts_fin_ms = str(int(dt_fin_aware.timestamp() * 1000))

    payload = {
        "queries": [{
            "refId": "A", "datasourceId": 5, "rawQuery": True, "resultFormat": "time_series",
            "query": f"SELECT last(\"value\")/1000 FROM \"NRG005\" WHERE (\"production_plant\" = 'medellin' AND \"variable_type\" = '{tag_energia}') AND $timeFilter GROUP BY time(1m) fill(null)"
        }],
        "range": {"from": iso_inicio, "to": iso_fin, "raw": {"from": iso_inicio, "to": iso_fin}},
        "from": ts_inicio_ms, "to": ts_fin_ms
    }
    try:
        resp = requests.post(URL_GRAFANA, json=payload, headers=HEADERS, auth=AUTH, verify=True, timeout=30)
        if resp.status_code == 200:
            return resp.json()
    except Exception as e:
        st.toast(f"Error HTTP en energía: {e}")
    return None

def parsear_energia(json_resp):
    df_A = extraer_dataframe_json(json_resp, 'A', 'Acumulador') if json_resp else pd.DataFrame(columns=['ts', 'Acumulador'])
    if df_A.empty:
        return pd.DataFrame(columns=['Fecha y hora', 'Energía [kWh]'])

    df_A['Energía [kWh]'] = df_A['Acumulador'].ffill().diff().round(4).fillna(0)
    df_A['Fecha y hora'] = pd.to_datetime(df_A['ts'], unit='ms').dt.tz_localize('UTC').dt.tz_convert('America/Bogota').dt.tz_localize(None).dt.floor('min')
    return df_A[['Fecha y hora', 'Energía [kWh]']]

def api_post_grafana_potencia(dt_inicio_aware, dt_fin_aware, tag_potencia):
    iso_inicio = dt_inicio_aware.astimezone(timezone.utc).strftime('%Y-%m-%dT%H:%M:%S.000Z')
    iso_fin = dt_fin_aware.astimezone(timezone.utc).strftime('%Y-%m-%dT%H:%M:%S.000Z')
    ts_inicio_ms = str(int(dt_inicio_aware.timestamp() * 1000))
    ts_fin_ms = str(int(dt_fin_aware.timestamp() * 1000))

    payload = {
        "queries": [{
            "refId": "A", "datasourceId": 5, "rawQuery": True, "resultFormat": "table",
            "query": f"SELECT last(\"value\")/1000 FROM \"NRG004\" WHERE (\"production_plant\" = 'medellin' AND \"variable_type\" = '{tag_potencia}') AND $timeFilter GROUP BY time(1s) fill(none)"
        }],
        "range": {"from": iso_inicio, "to": iso_fin, "raw": {"from": iso_inicio, "to": iso_fin}},
        "from": ts_inicio_ms, "to": ts_fin_ms
    }
    try:
        resp = requests.post(URL_GRAFANA, json=payload, headers=HEADERS, auth=AUTH, verify=True, timeout=30)
        if resp.status_code == 200:
            return resp.json()
    except Exception as e:
        st.toast(f"Error HTTP en potencia: {e}")
    return None

def parsear_potencia(json_resp):
    if not json_resp or 'A' not in json_resp.get('results', {}):
        return pd.DataFrame(columns=['Fecha y hora', 'Potencia [kW]'])
    try:
        tiempos = json_resp['results']['A']['frames'][0]['data']['values'][0]
        valores = json_resp['results']['A']['frames'][0]['data']['values'][1]
    except (KeyError, IndexError, TypeError):
        return pd.DataFrame(columns=['Fecha y hora', 'Potencia [kW]'])

    df = pd.DataFrame({'timestamp_ms': tiempos, 'Potencia [kW]': valores})
    if df.empty:
        return pd.DataFrame(columns=['Fecha y hora', 'Potencia [kW]'])

    df['Potencia [kW]'] = df['Potencia [kW]'].round(3)
    df['Fecha y hora'] = pd.to_datetime(df['timestamp_ms'], unit='ms').dt.tz_localize('UTC').dt.tz_convert('America/Bogota').dt.tz_localize(None).dt.floor('min')
    return df.groupby('Fecha y hora', as_index=False)['Potencia [kW]'].max()

# --- CONSULTA DE LA BASE DE DATOS EN DRIVE ---
st.divider()
with st.expander("🔎 Consultar base de datos maestra (Drive)", expanded=False):
    if st.button("📥 Cargar datos guardados en Drive", use_container_width=True):
        with st.spinner("Descargando energia.db desde Drive..."):
            try:
                drive_service = conectar_drive()
                file_id = buscar_db_en_drive(drive_service)

                if not file_id:
                    st.warning("Todavía no existe energia.db en la carpeta de Drive.")
                else:
                    descargar_db(drive_service, file_id)
                    conn = sqlite3.connect(RUTA_LOCAL_DB)
                    df_bd = pd.read_sql("SELECT * FROM registros_energia ORDER BY Timestamp DESC", conn)
                    conn.close()

                    st.success(f"✅ Conexión exitosa. La base tiene {len(df_bd)} registros en total.")

                    c1, c2, c3 = st.columns(3)
                    c1.metric("Registros totales", f"{len(df_bd):,}")
                    if not df_bd.empty:
                        c2.metric("Máquinas distintas", df_bd["ID_Maquina_Texto"].nunique())
                        c3.metric("Último dato", df_bd["Timestamp"].max())

                    st.markdown("**Últimos 50 registros guardados:**")
                    st.dataframe(df_bd.head(50), use_container_width=True)

                    st.session_state["df_bd_consulta"] = df_bd
            except Exception as e:
                st.error(f"Error al consultar Drive: {e}")

    if "df_bd_consulta" in st.session_state:
        df_bd = st.session_state["df_bd_consulta"]
        if not df_bd.empty:
            maquinas_bd = sorted(df_bd["ID_Maquina_Texto"].dropna().unique().tolist())
            maq_filtro = st.multiselect("Filtrar por máquina:", options=maquinas_bd, default=maquinas_bd, key="filtro_bd")
            df_filtrado = df_bd[df_bd["ID_Maquina_Texto"].isin(maq_filtro)]
            st.dataframe(df_filtrado, use_container_width=True)
            st.caption(f"Mostrando {len(df_filtrado)} de {len(df_bd)} registros totales.")

# --- INTERFAZ DE USUARIO ---
col1, col2 = st.columns(2)

with col1:
    st.markdown("#### ⏱️ Rango de Tiempo")
    c_f1, c_h1 = st.columns(2)
    d_inicio = c_f1.date_input("Fecha Inicio", value=pd.to_datetime("today") - pd.Timedelta(days=2))
    t_inicio = c_h1.time_input("Hora Inicio", value=pd.to_datetime("00:00").time())

    c_f2, c_h2 = st.columns(2)
    d_fin = c_f2.date_input("Fecha Fin", value=pd.to_datetime("today"))
    t_fin = c_h2.time_input("Hora Fin", value=pd.to_datetime("23:59").time())

with col2:
    st.markdown("#### ⚙️ Activos e Indicadores")
    maq_seleccionadas = st.multiselect("Selecciona Máquinas:", options=list(MAQUINAS_DISPONIBLES.keys()), default=["H80 (49)"])
    st.info("💡 La extracción trae **Energía y Potencia** automáticamente para cada máquina. Los valores negativos se corregirán a 0.")

dt_inicio_total = datetime.combine(d_inicio, t_inicio).replace(tzinfo=TZ_COLOMBIA)
dt_fin_total = datetime.combine(d_fin, t_fin).replace(tzinfo=TZ_COLOMBIA)

if st.button("🚀 Extraer Datos de InfluxDB", type="primary", use_container_width=True):
    if not maq_seleccionadas:
        st.error("⚠️ Debes seleccionar al menos una máquina.")
        st.stop()

    if dt_inicio_total >= dt_fin_total:
        st.error("⚠️ La fecha de inicio debe ser menor a la fecha de fin.")
        st.stop()

    ids_maquinas = [MAQUINAS_DISPONIBLES[m] for m in maq_seleccionadas]

    with st.status("Extrayendo datos desde InfluxDB...", expanded=True) as status:
        frames_encontrados = []

        total_dias = (dt_fin_total - dt_inicio_total).days + 1
        pasos_estimados = (total_dias // 5 + 1) * len(ids_maquinas)
        paso_actual = 0

        barra_progreso = st.progress(0)
        texto_progreso = st.empty()

        for id_maq in ids_maquinas:
            datos_maq = MAPEO_INFLUX[id_maq]
            nombre_maq = datos_maq["nombre"]
            tag_energia = datos_maq["tag_energia"]
            tag_potencia = datos_maq["tag_potencia"]

            ciclo_inicio = dt_inicio_total
            while ciclo_inicio < dt_fin_total:
                ciclo_fin = min(ciclo_inicio + timedelta(days=5), dt_fin_total)
                rango_str = f"{ciclo_inicio.strftime('%m-%d')} al {ciclo_fin.strftime('%m-%d')}"
                texto_progreso.write(f"🔄 Consultando: {nombre_maq} | {rango_str}")

                # 1. Extraemos Energía y Potencia
                json_energia = api_post_grafana_energia(ciclo_inicio, ciclo_fin, tag_energia)
                df_energia = parsear_energia(json_energia)

                json_potencia = api_post_grafana_potencia(ciclo_inicio, ciclo_fin, tag_potencia)
                df_potencia = parsear_potencia(json_potencia)

                # 2. Unimos ambas tablas por fecha y hora
                if not df_energia.empty or not df_potencia.empty:
                    df_tmp = df_energia.merge(df_potencia, on='Fecha y hora', how='outer')

                    df_tmp['Energía [kWh]'] = df_tmp['Energía [kWh]'].fillna(0)
                    df_tmp['Potencia [kW]'] = df_tmp['Potencia [kW]'].fillna(0)
                    df_tmp['maquina_o_puesto'] = nombre_maq
                    df_tmp['id_maquina_api'] = id_maq

                    frames_encontrados.append(df_tmp)

                ciclo_inicio = ciclo_fin + timedelta(minutes=1)
                paso_actual += 1
                barra_progreso.progress(min(paso_actual / pasos_estimados, 1.0))
                time.sleep(0.3)

        st.write("🗜️ Consolidando base de datos...")
        if frames_encontrados:
            df_final = pd.concat(frames_encontrados, ignore_index=True)
            df_final = df_final.sort_values(['Fecha y hora', 'maquina_o_puesto'])
            df_final = df_final.drop_duplicates(subset=["Fecha y hora", "maquina_o_puesto"], keep="first")

            if "Fecha y hora" in df_final.columns:
                df_final["Fecha y hora"] = pd.to_datetime(df_final["Fecha y hora"]).dt.strftime('%Y-%m-%d %H:%M:%S')

            # 3. Renombramos columnas para mantener compatibilidad
            renombres = {
                'Fecha y hora': 'Timestamp',
                'maquina_o_puesto': 'ID_Maquina_Texto',
                'Energía [kWh]': 'Energia_kWh',
                'Potencia [kW]': 'Potencia_kW'
            }
            df_final.rename(columns={k: v for k, v in renombres.items() if k in df_final.columns}, inplace=True)

            # 4. APLICAMOS EL CLIP (convertir negativos en 0) a las columnas correctas
            if "Energia_kWh" in df_final.columns:
                df_final["Energia_kWh"] = pd.to_numeric(df_final["Energia_kWh"], errors="coerce").clip(lower=0)
            if "Potencia_kW" in df_final.columns:
                df_final["Potencia_kW"] = pd.to_numeric(df_final["Potencia_kW"], errors="coerce").clip(lower=0)

            # Reordenar las columnas para mejor estética
            cols_orden = ["Timestamp", "ID_Maquina_Texto", "id_maquina_api", "Energia_kWh", "Potencia_kW"]
            df_final = df_final[[c for c in cols_orden if c in df_final.columns]]

            st.session_state['df_energia_extraido'] = df_final
            status.update(label=f"¡Extracción Exitosa! {len(df_final)} registros procesados.", state="complete", expanded=False)
        else:
            status.update(label="No se encontraron datos", state="error")
            st.error("🚩 No se encontraron datos para los parámetros ingresados.")

# --- RESULTADOS Y ACCIONES POST-EXTRACCIÓN ---
if 'df_energia_extraido' in st.session_state:
    df_mostrar = st.session_state['df_energia_extraido']
    st.success("✅ Datos listos para descarga, análisis o sincronización.")

    with st.expander("👀 Vista Previa de los Datos", expanded=True):
        st.dataframe(df_mostrar.head(100), use_container_width=True)

    # --- ANÁLISIS Y VISUALIZACIÓN ---
    def bloque_analisis(df_base: pd.DataFrame, variable_col: str, titulo: str, unidad: str, key_prefix: str):
        if variable_col not in df_base.columns:
            st.info(f"La columna {variable_col} no está disponible.")
            return

        maquinas_disponibles = sorted(df_base["ID_Maquina_Texto"].dropna().unique().tolist())

        col_a, col_b, col_c = st.columns(3)
        with col_a:
            granularidad = st.selectbox("Agrupar por:", options=["Minuto (sin agrupar)", "Hora", "Día"], key=f"{key_prefix}_gran")
        with col_b:
            estadistico = st.selectbox("Estadístico:", options=["Suma", "Promedio", "Mediana", "Mínimo", "Máximo", "Desviación estándar"], key=f"{key_prefix}_stat")
        with col_c:
            maquinas_sel = st.multiselect("Máquinas:", options=maquinas_disponibles, default=maquinas_disponibles, key=f"{key_prefix}_maq")

        df_analisis = df_base[df_base["ID_Maquina_Texto"].isin(maquinas_sel)].copy()
        df_analisis["Timestamp"] = pd.to_datetime(df_analisis["Timestamp"], errors="coerce")
        df_analisis[variable_col] = pd.to_numeric(df_analisis[variable_col], errors="coerce")
        df_analisis = df_analisis.dropna(subset=["Timestamp", variable_col])

        if df_analisis.empty:
            st.info(f"No hay datos de {titulo.lower()} para graficar.")
            return

        freq_map = {"Minuto (sin agrupar)": None, "Hora": "h", "Día": "D"}
        freq = freq_map[granularidad]

        func_map = {"Suma": "sum", "Promedio": "mean", "Mediana": "median", "Mínimo": "min", "Máximo": "max", "Desviación estándar": "std"}
        func = func_map[estadistico]

        df_analisis["Periodo"] = df_analisis["Timestamp"].dt.floor(freq) if freq else df_analisis["Timestamp"]

        tabla_resumen = df_analisis.groupby(["Periodo", "ID_Maquina_Texto"])[variable_col].agg(func).reset_index()

        st.markdown(f"**{estadistico} de {titulo} ({unidad}) por {granularidad.lower()}**")

        # Gráfico interactivo con Plotly
        fig = px.line(
            tabla_resumen,
            x="Periodo",
            y=variable_col,
            color="ID_Maquina_Texto",
            markers=True,
            template="plotly_white"
        )
        fig.update_layout(
            xaxis_title="Tiempo",
            yaxis_title=f"{titulo} ({unidad})",
            legend_title="Máquina",
            hovermode="x unified",
            margin=dict(l=0, r=0, t=30, b=0)
        )
        st.plotly_chart(fig, use_container_width=True)

        m1, m2, m3, m4 = st.columns(4)
        m1.metric("Suma total", f"{df_analisis[variable_col].sum():,.2f} {unidad}")
        m2.metric("Promedio", f"{df_analisis[variable_col].mean():,.2f} {unidad}")
        m3.metric("Mediana", f"{df_analisis[variable_col].median():,.2f} {unidad}")
        m4.metric("Máximo", f"{df_analisis[variable_col].max():,.2f} {unidad}")

    with st.expander("📊 Análisis y Visualización", expanded=True):
        if df_mostrar.empty:
            st.info("No hay datos para analizar.")
        else:
            tab_energia, tab_potencia = st.tabs(["⚡ Energía Consumida (kWh)", "🔌 Potencia (kW)"])
            with tab_energia:
                # Aquí se invoca el bloque para la columna Energia_kWh
                bloque_analisis(df_mostrar, "Energia_kWh", "Energía Consumida", "kWh", "energia")
            with tab_potencia:
                # Aquí se invoca el bloque para la columna Potencia_kW
                bloque_analisis(df_mostrar, "Potencia_kW", "Potencia", "kW", "potencia")

    col_dl, col_up = st.columns(2)

    with col_dl:
        csv_data = df_mostrar.to_csv(index=False, encoding='utf-8-sig')
        st.download_button(
            label="⬇️ Descargar como CSV",
            data=csv_data,
            file_name=f"Kern_Influx_Energia_{datetime.now().strftime('%Y%m%d_%H%M')}.csv",
            mime="text/csv",
            use_container_width=True
        )

    # --- DESCARGA POR MÁQUINA (SOBRE LA PLANTILLA OFICIAL) ---
    st.divider()
    st.markdown("#### 📦 Descarga por máquina (formato plantilla)")
    st.caption(
        "Genera un archivo independiente por cada máquina consultada, con solo dos columnas: "
        "**Demanda [kW]** (= Potencia_kW) y **Fecha en UTC** — igual que `Plantilla_Energia-UTC.xlsx`."
    )

    col_fmt, col_gen = st.columns([1, 2])
    with col_fmt:
        formato_export_maq = st.radio(
            "Formato:", ["CSV", "Excel (plantilla)"], horizontal=True, key="formato_export_maq"
        )

    with col_gen:
        if st.button("📦 Generar archivos por máquina", use_container_width=True):
            df_export = df_mostrar.copy()
            df_export["Fecha_UTC"] = df_export["Timestamp"].apply(_timestamp_bogota_a_utc_iso)
            df_export["Fecha_UTC_dt"] = df_export["Timestamp"].apply(_timestamp_bogota_a_utc_dt)

            zip_buffer = BytesIO()
            with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
                for nombre_maq, df_maq in df_export.groupby("ID_Maquina_Texto"):
                    df_maq = df_maq.sort_values("Timestamp")
                    if formato_export_maq == "CSV":
                        zf.writestr(f"{nombre_maq}.csv", generar_csv_maquina(df_maq))
                    else:
                        zf.writestr(f"{nombre_maq}.xlsx", generar_xlsx_maquina(df_maq))

            zip_buffer.seek(0)
            n_maquinas = df_export["ID_Maquina_Texto"].nunique()
            st.session_state["zip_por_maquina"] = zip_buffer.getvalue()
            st.session_state["zip_por_maquina_info"] = (formato_export_maq, n_maquinas)

        if "zip_por_maquina" in st.session_state:
            formato_guardado, n_maquinas = st.session_state["zip_por_maquina_info"]
            st.download_button(
                label=f"⬇️ Descargar ZIP ({formato_guardado}) - {n_maquinas} máquina(s)",
                data=st.session_state["zip_por_maquina"],
                file_name=f"Energia_por_maquina_{datetime.now().strftime('%Y%m%d_%H%M')}.zip",
                mime="application/zip",
                use_container_width=True
            )

    with col_up:
        if st.button("☁️ Sincronizar a Google Drive (SQLite)", type="primary", use_container_width=True):
            with st.spinner("Sincronizando con la base de datos maestra en Drive..."):
                try:
                    drive_service = conectar_drive()

                    # 1. Buscar y descargar el .db existente (si ya hay uno en la carpeta)
                    file_id = buscar_db_en_drive(drive_service)
                    if file_id:
                        descargar_db(drive_service, file_id)
                    elif os.path.exists(RUTA_LOCAL_DB):
                        os.remove(RUTA_LOCAL_DB)  # aseguramos que no quede un .db viejo de otra sesión

                    # 2. Abrir/crear la tabla y hacer el upsert
                    conn = sqlite3.connect(RUTA_LOCAL_DB)
                    preparar_tabla(conn)

                    df_para_db = df_mostrar.copy()
                    df_para_db["Timestamp"] = df_para_db["Timestamp"].astype(str)
                    df_para_db["id_maquina_api"] = df_para_db["id_maquina_api"].astype(int)

                    antes = conn.execute("SELECT COUNT(*) FROM registros_energia").fetchone()[0]
                    upsertar_dataframe(conn, df_para_db)
                    despues = conn.execute("SELECT COUNT(*) FROM registros_energia").fetchone()[0]
                    conn.close()

                    # 3. Subir el .db actualizado de vuelta a Drive
                    subir_db(drive_service, file_id)

                    nuevos = despues - antes
                    if nuevos > 0:
                        st.success(f"🎉 ¡Éxito! Se agregaron {nuevos} registros nuevos. Total en base: {despues}.")
                    else:
                        st.info(f"👍 Todo al día. Los {len(df_para_db)} registros extraídos ya existían en la base. Total: {despues}.")

                except Exception as e:
                    st.error(f"Error de conexión con Drive: {e}")
